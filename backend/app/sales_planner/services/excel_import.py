"""Parses the Campana Plan.xlsx workbook into the sales planner DB.

The workbook ships with four sheets:
  - "General"   — KPI dashboard per BC.
  - "Plan CP"   — campaign tasks per location/date.
  - "PR Grupo"  — promoter staff with daily allocation columns.
  - "Ubicacion" — location master data.

The parser is intentionally tolerant: headers are detected by content rather
than position, Excel serial dates are converted, and warnings are emitted for
unmappable rows rather than aborting the whole import.
"""
from __future__ import annotations

import hashlib
import io
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from typing import Iterable

import openpyxl
from openpyxl.cell import Cell
from sqlalchemy.orm import Session

from app.sales_planner.models import (
    Branch, BusinessCenter, CampaignTask, CampaignTarget, DailyPlan,
    ImportedFile, Location, Merchandising, PRGroup, PRStaff,
    SalesAuditLog, SalesCampaign,
)


# ============================================================
# Generic helpers
# ============================================================

EXCEL_EPOCH = datetime(1899, 12, 30)  # Excel's day 1 corresponds to 1900-01-01


def excel_serial_to_date(value) -> date | None:
    """Convert an Excel serial number (or already-parsed date) to a Python date."""
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, (int, float)):
        try:
            return (EXCEL_EPOCH + timedelta(days=float(value))).date()
        except (OverflowError, ValueError):
            return None
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d/%m/%y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value, fmt).date()
            except ValueError:
                continue
    return None


def to_int(value, default=0) -> int:
    if value is None or value == "":
        return default
    try:
        return int(float(value))
    except (TypeError, ValueError):
        return default


def to_float(value, default=0.0) -> float:
    if value is None or value == "":
        return default
    try:
        return float(value)
    except (TypeError, ValueError):
        return default


def norm_str(value) -> str | None:
    if value is None:
        return None
    s = str(value).strip()
    return s or None


# ============================================================
# Result containers
# ============================================================

@dataclass
class ImportResult:
    file_id: int
    sheets_detected: list[str]
    rows_per_sheet: dict[str, int] = field(default_factory=dict)
    inserted: dict[str, int] = field(default_factory=dict)
    warnings: list[str] = field(default_factory=list)
    errors: list[str] = field(default_factory=list)


# ============================================================
# Importer
# ============================================================

class CampanaPlanImporter:
    """Parses and commits a single Campana Plan workbook.

    Usage:
        importer = CampanaPlanImporter(db)
        result = importer.import_bytes(payload, filename="plan.xlsx", uploaded_by=user.id)
    """

    EXPECTED_SHEETS = {"General", "Plan CP", "PR Grupo", "Ubicacion"}

    def __init__(self, db: Session) -> None:
        self.db = db
        self.result = ImportResult(file_id=0, sheets_detected=[])
        self._branches_cache: dict[str, Branch] = {}
        self._bcs_cache: dict[str, BusinessCenter] = {}
        self._groups_cache: dict[str, PRGroup] = {}
        self._locations_cache: dict[str, Location] = {}
        self._staff_cache: dict[str, PRStaff] = {}

    # -------- public entry --------

    def import_bytes(
        self, payload: bytes, *, filename: str, uploaded_by: int | None = None,
        commit: bool = True,
    ) -> ImportResult:
        sha = hashlib.sha256(payload).hexdigest()
        existing = (
            self.db.query(ImportedFile).filter(ImportedFile.sha256 == sha).first()
        )
        if existing and existing.status == "committed":
            self.result.warnings.append(
                f"Identical file already imported as id={existing.id}"
            )
            self.result.file_id = existing.id
            return self.result

        imported = ImportedFile(
            filename=filename, original_name=filename, sha256=sha,
            size_bytes=len(payload), uploaded_by=uploaded_by, status="parsed",
        )
        self.db.add(imported)
        self.db.flush()
        self.result.file_id = imported.id

        wb = openpyxl.load_workbook(io.BytesIO(payload), data_only=True)
        self.result.sheets_detected = list(wb.sheetnames)

        # Order matters: master → staff → tasks.
        if "Ubicacion" in wb.sheetnames:
            self._parse_ubicacion(wb["Ubicacion"])
        if "PR Grupo" in wb.sheetnames:
            self._parse_pr_grupo(wb["PR Grupo"])
        if "Plan CP" in wb.sheetnames:
            self._parse_plan_cp(wb["Plan CP"])
        # General sheet only used for dashboard pre-population (no critical writes).
        if "General" in wb.sheetnames:
            self._parse_general(wb["General"])

        if commit:
            imported.status = "committed"
            imported.sheets_detected = len(self.result.sheets_detected)
            imported.rows_imported = sum(self.result.inserted.values())
            imported.committed_at = datetime.utcnow()
            self._audit("import", imported.id, after={"sha": sha})
            self.db.commit()
        else:
            self.db.flush()

        return self.result

    # -------- sheet parsers --------

    def _parse_ubicacion(self, ws) -> None:
        """Row 2 holds headers in `Campana Plan.xlsx`. Row 3+ holds data.

        Required headers (case-insensitive trim):
          Code Ubicacion / BR / BC / DEPARTAMENTO / DISTRITO / Tipo DF/CP /
          Horario traffico / Fecha Alta Traffico / Prioridad / Latitud / Longitud
        """
        headers = self._read_headers(ws, header_row=2)
        if not headers:
            self.result.warnings.append("Ubicacion: header row missing")
            return

        sheet_inserted = 0
        for row in ws.iter_rows(min_row=3, values_only=True):
            code = norm_str(self._cell(row, headers, "Code Ubicacion"))
            if not code or not code.startswith("CP"):
                continue
            br_code = norm_str(self._cell(row, headers, "BR"))
            bc_code = norm_str(self._cell(row, headers, "BC"))
            if not br_code or not bc_code:
                self.result.warnings.append(f"Ubicacion {code}: missing BR/BC")
                continue
            branch = self._get_or_create_branch(br_code)
            bc = self._get_or_create_bc(bc_code, branch)
            loc = self._locations_cache.get(code)
            if loc is None:
                loc = Location(code=code, branch_id=branch.id, business_center_id=bc.id)
                self.db.add(loc)
                self._locations_cache[code] = loc
            loc.departamento = norm_str(self._cell(row, headers, "DEPARTAMENTO"))
            loc.distrito = norm_str(self._cell(row, headers, "DISTRITO"))
            loc.tipo_df_cp = norm_str(self._cell(row, headers, "Tipo DF/CP"))
            loc.horario_traffico = norm_str(self._cell(row, headers, "Horario traffico"))
            loc.fecha_alta_traffico = norm_str(self._cell(row, headers, "Fecha Alta Traffico"))
            loc.prioridad = to_int(self._cell(row, headers, "Prioridad"), 1)
            loc.latitud = to_float(self._cell(row, headers, "Latitud"))
            loc.longitud = to_float(self._cell(row, headers, "Longitud"))
            loc.nota = norm_str(self._cell(row, headers, "Nota"))
            # meta defaults
            loc.meta_prepago = to_int(self._cell(row, headers, "Prepago"))
            loc.meta_postpago = to_int(self._cell(row, headers, "Postpago"))
            loc.meta_bipay = to_int(self._cell(row, headers, "Bipay cashin 20%\n(minimo 5 soles)"))
            loc.meta_tv360 = to_int(self._cell(row, headers, "TV360"))
            loc.meta_mnp = to_int(self._cell(row, headers, "MNP"))
            loc.meta_agentes = to_int(self._cell(row, headers, "Creacion de Agentes/Negocio"))
            loc.meta_usuarios_bipay = to_int(self._cell(row, headers, "Usuarios Bipay"))
            loc.meta_pago_servicios = to_int(self._cell(row, headers, "Pago de Servicios"))
            loc.meta_tusami = to_int(self._cell(row, headers, "Meta Tusami"))
            # gasto
            loc.gasto_comida = to_float(self._cell(row, headers, "Pago Comida"))
            loc.gasto_hotel = to_float(self._cell(row, headers, "Pago Hotel"))
            loc.gasto_movilidad = to_float(self._cell(row, headers, "Pago Movilidad"))
            loc.gasto_renta_local = to_float(self._cell(row, headers, "Pago Renta Local"))
            # merch
            loc.merch_boligrafo = to_int(self._cell(row, headers, "Bolígrafo"))
            loc.merch_taza = to_int(self._cell(row, headers, "Taza"))
            loc.merch_llavero = to_int(self._cell(row, headers, "LLavero"))
            loc.merch_papin = to_int(self._cell(row, headers, "Papin"))
            loc.merch_sombrero = to_int(self._cell(row, headers, "Sombrero"))
            sheet_inserted += 1
        self.db.flush()
        self.result.rows_per_sheet["Ubicacion"] = sheet_inserted
        self.result.inserted["locations"] = sheet_inserted

    def _parse_pr_grupo(self, ws) -> None:
        """Row 3 holds headers; daily allocation columns are themselves dates."""
        headers = self._read_headers(ws, header_row=3)
        if not headers:
            self.result.warnings.append("PR Grupo: header row missing")
            return

        # Detect daily-allocation columns: header cell is a date.
        day_cols: dict[int, date] = {}
        for idx, h in enumerate(headers, start=1):
            d = excel_serial_to_date(h)
            if d is not None:
                day_cols[idx] = d

        sheet_inserted = 0
        for row in ws.iter_rows(min_row=4, values_only=True):
            pr_code = norm_str(self._cell(row, headers, "PR Code"))
            if not pr_code:
                continue
            br_code = norm_str(self._cell(row, headers, "BR"))
            bc_code = norm_str(self._cell(row, headers, "BC"))
            if not br_code or not bc_code:
                self.result.warnings.append(f"PR {pr_code}: missing BR/BC")
                continue
            branch = self._get_or_create_branch(br_code)
            bc = self._get_or_create_bc(bc_code, branch)
            group_code = norm_str(self._cell(row, headers, "Grupo"))
            group = self._get_or_create_group(group_code, bc) if group_code else None

            staff = self._staff_cache.get(pr_code)
            if staff is None:
                staff = PRStaff(
                    pr_code=pr_code, branch_id=branch.id,
                    business_center_id=bc.id,
                )
                self.db.add(staff)
                self._staff_cache[pr_code] = staff
            staff.owner_code = norm_str(self._cell(row, headers, "owner_code"))
            staff.tipo_pr = norm_str(self._cell(row, headers, "Tipo de PR"))
            staff.leader = norm_str(self._cell(row, headers, "Leader (AF/BC)"))
            staff.kpi_trabajo = to_int(self._cell(row, headers, "KPI Trabajo"))
            staff.cantidad_dia_trabajo = to_int(
                self._cell(row, headers, "Cantidad Dia de Trabajo"), 18,
            )
            staff.estado = norm_str(self._cell(row, headers, "OK/NO OK")) or "OK"
            if group is not None:
                staff.group_id = group.id
            self.db.flush()

            # Daily allocations
            for col_idx, plan_date in day_cols.items():
                units = to_float(row[col_idx - 1] if col_idx - 1 < len(row) else None)
                if units > 0:
                    dp = DailyPlan(
                        pr_staff_id=staff.id, group_id=group.id if group else None,
                        plan_date=plan_date, units=min(units, 1.0),
                    )
                    self.db.add(dp)
            sheet_inserted += 1

        self.db.flush()
        self.result.rows_per_sheet["PR Grupo"] = sheet_inserted
        self.result.inserted["pr_staff"] = sheet_inserted

    def _parse_plan_cp(self, ws) -> None:
        """Row 6 holds headers; row 7+ rows; columns from 66 onward are calendar dates."""
        headers = self._read_headers(ws, header_row=6)
        if not headers:
            self.result.warnings.append("Plan CP: header row missing")
            return

        # The Excel reader returns date-headers as datetime instances.
        day_cols: dict[int, date] = {}
        for idx, h in enumerate(headers, start=1):
            d = excel_serial_to_date(h)
            if d is not None:
                day_cols[idx] = d

        # Auto-create a campaign for this import.
        first_date = min(day_cols.values()) if day_cols else date.today()
        last_date = max(day_cols.values()) if day_cols else first_date
        camp = SalesCampaign(
            name=f"Campaign — {first_date.isoformat()}",
            start_date=first_date,
            end_date=last_date,
            horizon_days=(last_date - first_date).days + 1,
        )
        self.db.add(camp)
        self.db.flush()

        sheet_inserted = 0
        for row in ws.iter_rows(min_row=7, values_only=True):
            code = norm_str(self._cell(row, headers, "Code Ubicacion"))
            if not code or not code.startswith("CP"):
                continue
            location = self._locations_cache.get(code)
            if location is None:
                self.result.warnings.append(f"Plan CP {code}: unknown location, skipping")
                continue
            task = CampaignTask(
                campaign_id=camp.id,
                task_name=f"{location.distrito or code} — {location.tipo_df_cp or 'Campaign'}",
                location_id=location.id,
                business_center_id=location.business_center_id,
                distrito=location.distrito,
                tipo_df_cp=location.tipo_df_cp,
                horario_traffico=location.horario_traffico,
                fecha_alta_traffico=location.fecha_alta_traffico,
                people_in_charge=norm_str(self._cell(row, headers, "People InCharge")),
                start_date=excel_serial_to_date(self._cell(row, headers, "Start Date")),
                end_date=excel_serial_to_date(self._cell(row, headers, "End")),
                priority=location.prioridad,
                df_bts_code=norm_str(self._cell(row, headers, "DFCode/BTS Code")),
                notes=norm_str(self._cell(row, headers, "Nota")),
                cumple_activaciones=self._bool(self._cell(row, headers, "CUMPLE ACTIVACIONES")),
                cumple_digital=self._bool(self._cell(row, headers, "Cumple Digital ")),
                campana_ok=norm_str(self._cell(row, headers, "CAMPAÑA OK")),
            )
            if task.start_date and task.end_date:
                task.duration_days = (task.end_date - task.start_date).days + 1
            # group/PR
            grupo_code = norm_str(self._cell(row, headers, "Grupo/Code PR"))
            if grupo_code:
                group = self._get_or_create_group(grupo_code, None)
                task.group_id = group.id
            self.db.add(task)
            self.db.flush()

            # Targets from row meta
            tgt = CampaignTarget(
                task_id=task.id,
                prepago=to_int(self._cell(row, headers, "Prepago")),
                postpago=to_int(self._cell(row, headers, "Postpago")),
                bipay=to_int(self._cell(row, headers, "Bipay cashin 20%\n(minimo 5 soles)")),
                tv360=to_int(self._cell(row, headers, "TV360")),
                mnp=to_int(self._cell(row, headers, "MNP")),
                agentes=to_int(self._cell(row, headers, "Creacion de Agentes/Negocio")),
                usuarios_bipay=to_int(self._cell(row, headers, "Usuarios Bipay")),
                pago_servicios=to_int(self._cell(row, headers, "Pago de Servicios")),
                tusami=to_int(self._cell(row, headers, "Meta Tusami")),
            )
            self.db.add(tgt)
            # Merch
            merch = Merchandising(
                task_id=task.id,
                boligrafo=to_int(self._cell(row, headers, "Bolígrafo")),
                taza=to_int(self._cell(row, headers, "Taza")),
                llavero=to_int(self._cell(row, headers, "LLavero")),
                papin=to_int(self._cell(row, headers, "Papin")),
                sombrero=to_int(self._cell(row, headers, "Sombrero")),
            )
            self.db.add(merch)

            # Daily plan for each date column with a value
            for col_idx, plan_date in day_cols.items():
                val = row[col_idx - 1] if col_idx - 1 < len(row) else None
                if val:
                    dp = DailyPlan(
                        task_id=task.id, plan_date=plan_date,
                        units=min(to_float(val, 1.0), 1.0),
                    )
                    self.db.add(dp)
            sheet_inserted += 1

        self.db.flush()
        self.result.rows_per_sheet["Plan CP"] = sheet_inserted
        self.result.inserted["campaign_tasks"] = sheet_inserted

    def _parse_general(self, ws) -> None:
        """Pre-populate dashboard cache only (no critical writes); future iteration."""
        self.result.rows_per_sheet["General"] = ws.max_row

    # -------- helpers --------

    def _get_or_create_branch(self, code: str) -> Branch:
        if code in self._branches_cache:
            return self._branches_cache[code]
        branch = self.db.query(Branch).filter(Branch.code == code).first()
        if not branch:
            branch = Branch(code=code)
            self.db.add(branch)
            self.db.flush()
        self._branches_cache[code] = branch
        return branch

    def _get_or_create_bc(self, code: str, branch: Branch) -> BusinessCenter:
        key = code
        if key in self._bcs_cache:
            return self._bcs_cache[key]
        bc = self.db.query(BusinessCenter).filter(BusinessCenter.code == code).first()
        if not bc:
            bc = BusinessCenter(code=code, branch_id=branch.id)
            self.db.add(bc)
            self.db.flush()
        self._bcs_cache[key] = bc
        return bc

    def _get_or_create_group(self, code: str, bc: BusinessCenter | None) -> PRGroup:
        if code in self._groups_cache:
            return self._groups_cache[code]
        group = self.db.query(PRGroup).filter(PRGroup.code == code).first()
        if not group:
            group = PRGroup(
                code=code,
                business_center_id=bc.id if bc else None,
            )
            self.db.add(group)
            self.db.flush()
        self._groups_cache[code] = group
        return group

    @staticmethod
    def _read_headers(ws, header_row: int) -> list:
        return [c.value for c in ws[header_row]]

    @staticmethod
    def _cell(row: tuple, headers: list, column_name: str):
        """Return the cell value matching `column_name` in `headers` (case/space insensitive)."""
        target = column_name.strip().lower()
        for i, h in enumerate(headers):
            if h is None:
                continue
            if str(h).strip().lower() == target:
                return row[i] if i < len(row) else None
        return None

    @staticmethod
    def _bool(value) -> bool | None:
        if value is None or value == "":
            return None
        s = str(value).strip().lower()
        if s in ("ok", "true", "1", "yes", "si"):
            return True
        if s in ("no ok", "false", "0", "no"):
            return False
        return None

    def _audit(self, action: str, entity_id: int, after: dict) -> None:
        import json
        self.db.add(SalesAuditLog(
            user_id=None, entity="imported_file", entity_id=entity_id,
            action=action, after_state=json.dumps(after, default=str),
        ))
