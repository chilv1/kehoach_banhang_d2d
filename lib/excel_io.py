"""Đọc/ghi file Excel cho hệ thống quản lý campaign."""
from __future__ import annotations

import io

import openpyxl
import pandas as pd

from . import campaign, db


# ---------- Import master data ----------

def import_master_xlsx(file_like) -> dict:
    """Đọc Excel mẫu Campana Plan.xlsx và import 2 sheet master:
    `Ubicacion` và `PR Grupo`. Không tạo campaign ở bước này — dùng
    Plan Generator để sinh DRAFT campaign sau khi import master.
    """
    wb = openpyxl.load_workbook(file_like, data_only=True)
    counts = {"ubicacion": 0, "promoter": 0}

    if "Ubicacion" in wb.sheetnames:
        counts["ubicacion"] = _import_ubicacion(wb["Ubicacion"])
    if "PR Grupo" in wb.sheetnames:
        counts["promoter"] = _import_promoter(wb["PR Grupo"])
    return counts


def _cell(row, idx):
    if idx < 1 or idx > len(row):
        return None
    return row[idx - 1]


def _import_ubicacion(ws) -> int:
    rows = []
    for r in ws.iter_rows(min_row=3, values_only=True):
        code = _cell(r, 1)
        if not code or not isinstance(code, str) or not str(code).strip():
            continue
        row = {
            "code": str(code).strip(),
            "br": str(_cell(r, 2) or "").strip(),
            "bc": str(_cell(r, 3) or "").strip(),
            "departamento": _cell(r, 4),
            "distrito": _cell(r, 5),
            "tipo_dfcp": _cell(r, 6),
            "horario_traffico": _cell(r, 7),
            "fecha_alta_traffico": _cell(r, 8),
            "prioridad": int(_cell(r, 9) or 1),
            "latitud": _num(_cell(r, 10)),
            "longitud": _num(_cell(r, 11)),
            "nota": _cell(r, 12),
            "meta_prepago": _num(_cell(r, 13)),
            "meta_postpago": _num(_cell(r, 14)),
            "meta_bipay": _num(_cell(r, 15)),
            "meta_tv360": _num(_cell(r, 16)),
            "meta_mnp": _num(_cell(r, 17)),
            "meta_agentes": _num(_cell(r, 18)),
            "meta_usuarios_bipay": _num(_cell(r, 19)),
            "meta_pago_servicios": _num(_cell(r, 20)),
            "meta_tusami": _num(_cell(r, 21)),
            "gasto_comida": _num(_cell(r, 22)),
            "gasto_hotel": _num(_cell(r, 23)),
            "gasto_movilidad": _num(_cell(r, 24)),
            "gasto_renta": _num(_cell(r, 25)),
            "merch_boligrafo": _int(_cell(r, 27)),
            "merch_taza": _int(_cell(r, 28)),
            "merch_llavero": _int(_cell(r, 29)),
            "merch_papin": _int(_cell(r, 30)),
            "merch_sombrero": _int(_cell(r, 31)),
            "cantidad_dia": 1,
            "activo": 1,
        }
        if not row["br"] or not row["bc"]:
            continue
        rows.append(row)
    return db.bulk_upsert_ubicacion(rows)


def _import_promoter(ws) -> int:
    rows = []
    for r in ws.iter_rows(min_row=4, values_only=True):
        pr_code = _cell(r, 3)
        if not pr_code:
            continue
        row = {
            "pr_code": str(pr_code).strip(),
            "br": str(_cell(r, 1) or "").strip(),
            "bc": str(_cell(r, 2) or "").strip(),
            "owner_code": _cell(r, 4),
            "tipo": _cell(r, 5),
            "leader": _cell(r, 6),
            "grupo": _cell(r, 7),
            "kpi_trabajo": _num(_cell(r, 8)),
            "cantidad_dia_trabajo": _int(_cell(r, 9), default=18),
            "estado": _cell(r, 10) or "OK",
            "activo": 1,
        }
        if not row["br"] or not row["bc"]:
            continue
        rows.append(row)
    return db.bulk_upsert_promoter(rows)


def _num(v) -> float:
    if v is None or v == "":
        return 0.0
    try:
        return float(v)
    except (ValueError, TypeError):
        return 0.0


def _int(v, default: int = 0) -> int:
    if v is None or v == "":
        return default
    try:
        return int(float(v))
    except (ValueError, TypeError):
        return default


# ---------- Export ----------

def export_campaigns_xlsx(month: str) -> bytes:
    """Xuất campaign + result của tháng (YYYY-MM) ra Excel nhiều sheet."""
    camps = campaign.list_campaigns(month=month)
    results = campaign.list_results_with_campaign(month=month)
    ubic = db.list_ubicacion()
    prs = db.list_promoter()

    buf = io.BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        if not camps.empty:
            camps.to_excel(writer, sheet_name="Campaigns", index=False)
            pivot = camps.pivot_table(
                index=["bc", "ubicacion_code"],
                columns="fecha",
                values="status",
                aggfunc="first",
            ).fillna("")
            pivot.to_excel(writer, sheet_name="Calendar")
        else:
            pd.DataFrame({"info": ["Chưa có campaign"]}).to_excel(
                writer, sheet_name="Campaigns", index=False
            )
        if not results.empty:
            results.to_excel(writer, sheet_name="Results", index=False)
        ubic.to_excel(writer, sheet_name="Ubicacion", index=False)
        prs.to_excel(writer, sheet_name="PR Grupo", index=False)
    buf.seek(0)
    return buf.getvalue()
